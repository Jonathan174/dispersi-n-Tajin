import os
import shutil
import openpyxl
import time

from openpyxl import styles
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from fileSelection import seleccionar_archivo, filtrar_datos, calcular_porcentaje

def parent_of_merged_cell(cell, sheet):
    """ Find the parent of the merged cell by iterating through the range of merged cells """
    sheet = cell.parent
    child_coord = cell.coordinate

    # Note: if there are many merged cells in a large spreadsheet, this may become inefficient
    for merged in sheet.merged_cells.ranges:
        if child_coord in merged:
            return merged.start_cell.coordinate
    return None

def merged_value(cell, sheet):
    """ Reads the value of a cell, if cell is within a merged cell,
        find the first cell in the merged cell and get its value
    """
    if isinstance(cell, openpyxl.cell.cell.Cell):
        return cell.value
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        coord = parent_of_merged_cell(cell, sheet)
        parent = sheet[coord]
        return parent.value

def format_file(file_name: str, color_range: int, family_colors: tuple) -> None: 
    
    wb1 = load_workbook(filename=file_name) 
    sheets = wb1.sheetnames
    sheet = sheets[0] # La primera y única hoja, en caso de que haya más hojas se puede ajustar por índice o por nombre
    ws1 = wb1[sheet] 
    ws1.sheet_view.showGridLines = False



    # Definir el estilo de borde grueso
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    
    # Guarda el rango de celdas del total de cada familia, servirá para verificar hasta dónde va a agrupar los encabezados de las filas 2 y 3
    dv = DataValidation(type="list", formula1='"Opción1,Opción2,Opción3"', allow_blank=True) 

    last_title = 0 # Guarda la columna en la que se encuentra el último encabezado de la fila 1
    last_subtitle = 0   # Guarda la columna en la que está el último encabezado de la fila 2 (la 2 y la 3 tienen el mismo tamaño, solo revisa la 2)

    max_length = 0
    column_b = ws1['B']  # Columna B
    for cell in column_b:
        if cell.value:  # Si la celda tiene contenido
            # Estima la longitud del texto en la celda. 
            # Puedes ajustar el factor multiplicativo según necesites para mejor ajuste visual
            text_length = len(str(cell.value))
            max_length = max(max_length, text_length)

    # Ajustar el ancho de la columna B
    ws1.column_dimensions[get_column_letter(2)].width = max_length

    indice =0
    for row in ws1: 
        start_cell = None   # Aquí va a guardar la primera celda en la que encuentre un valor, se va a agrupar hasta donde encuentre end_cell
        end_cell = None # Cuando encuentre la siguiente celda con valor en la misma fila, guarda aquí esa celda

        for cell in row:
            cell.alignment = styles.Alignment(horizontal='center', vertical='center')   # Centrar el contenido de todas las celdas
            
            if cell.row in (1,2,3): # En las primeras 3 filas es donde están las celdas a combinar (encabezados)
                if cell.value is not None:
                    if start_cell is None:  # El primer encabezado de esa fila
                        start_cell = cell
                    if cell.value != start_cell.value: # Busca la siguiente celda con valor en esa fila
                        end_cell = cell
                else:
                    if start_cell is not None:
                        if start_cell.row == 1 and start_cell.column > last_title: # Condición para verificar el último encabezado de la fila 1
                            last_title = start_cell.column
                        elif start_cell.row ==2 and start_cell.column > last_subtitle: # Condición para verificar el último encabezado de la fila 2
                            last_subtitle = start_cell.column
                    if start_cell is not None and end_cell is not None: # Si ya tiene definidos inicio y fin del encabezado, va a combinar celdas
                        if start_cell.row == 1:

                            # Condición que verifica si el encabezado es el total de la familia, para agrupar las 9 celdas
                            if ws1.cell(2, start_cell.column).value is None: 
                                # Guarda las coordenadas de la última celda de la combinación del total de cada familia
                                temp_end_cell = ws1.cell(row=end_cell.row+2, column=end_cell.column-1).coordinate

                                # Agrupa hasta una columna antes del siguiente encabezado y 2 filas abajo
                                ws1.merge_cells(start_row=start_cell.row, start_column=start_cell.column, end_row=end_cell.row+2, end_column=end_cell.column-1)
                                dv.add(f"{start_cell.coordinate}:{temp_end_cell}") # guarda el rango combinado de las 9 celdas del total de cada familia
                            else:
                                # Condición que verifica si el encabezado solo es un encabezado que combina en la misma fila
                                ws1.merge_cells(start_row=start_cell.row, start_column=start_cell.column, end_row=end_cell.row, end_column=end_cell.column-1)
                        
                        if start_cell.row in (2,3):
                            # Encuentra un encabezado y el siguiente, si la celda anterior al siguiente encabezado no es parte del total, agrupa hasta esa celda
                            if ws1.cell(end_cell.row, end_cell.column-1) not in dv:
                                ws1.merge_cells(start_row=start_cell.row, start_column=start_cell.column, end_row=end_cell.row, end_column=end_cell.column-1)

                            # Si la celda anterior al siguiente encabezado es parte del total, agrupa con la cuarta celda antes del encabezado (antes de las 3 del total)
                            elif ws1.cell(end_cell.row, end_cell.column-1) in dv:
                                ws1.merge_cells(start_row=start_cell.row, start_column=start_cell.column, end_row=end_cell.row, end_column=end_cell.column-4)

                        start_cell = end_cell # El siguiente encabezado se convierte en el primero, ahora va a verificar desde ese en adelante
                        end_cell = None
                
                # Revisa el valor de la primera fila de cada columna para asignarle el color correspondiente a cada familia, 
                # se necesita la función porque ya están combinadas las celdas
                if merged_value(ws1.cell(1, cell.column), ws1) is not None: 
                    # Hay encabezados que tienen más de una palabra, separa el encabezado por palabras y revisa la primera (solo la primera, por los que solo tienen una palabra)
                    family = merged_value(ws1.cell(1, cell.column), ws1).split()
                    if family[0] == "Chile": family[0] = family[0] + family[1] # Tiene 2 palabras, solo verifica la primera y sobreescribe con el valor de ambas
                    
                    # La primera palabra del encabezado es la key para el diccionario de colores
                    cell.fill = styles.PatternFill(start_color=family_colors[family[0]], end_color=family_colors[family[0]], fill_type='solid')

                if cell.row == 2 and cell.column == 1: # Verifica que haya terminado de leer la primera fila para agrupar el último encabezado
                    ws1.merge_cells(start_row=1, start_column=last_title, end_row=3, end_column=last_title+2)
                
                if cell.column > 2:
                    cell.border = thick_border  # Aplica el borde grueso
            
            if cell.row == 4 and cell.column == 1: # Verifica que haya terminado de leer las filas 2 y 3 para agrupar el último encabezado respectivamente
                ws1.merge_cells(start_row=2, start_column=last_subtitle, end_row=2, end_column=last_subtitle+2)
                ws1.merge_cells(start_row=3, start_column=last_subtitle, end_row=3, end_column=last_subtitle+2)
                
                # Aplicar borde grueso a toda la fila 4
                for cell in ws1[4]:  # ws1[4] accede a todas las celdas de la fila 4
                    cell.border = thick_border

            elif cell.row >=5:
                
                if ws1.cell(4, cell.column).value == '%': # Todas las demás columnas ya son de valores, verifica que la columna sea la de porcentaje
                    if cell.value is not None and isinstance(cell.value, (int, float)): # Revisa que el valor de la celda sea un número para asignarle color de fuente
                        cell.number_format = '0%'
                        if cell.value >= 1:
                            cell.font = styles.Font(color="27961c") # green color
                        elif cell.value > color_range and cell.value < 1:
                            cell.font = styles.Font(color="0000ff") # blue color
                        elif cell.value <= color_range:
                            cell.font = styles.Font(color="ff0000") # red color

                # Condición 1: Si estás en la columna A y el valor no es un dígito, pinta toda la fila de gris
                if cell.column == 1 and cell.value is not None:  # openpyxl las columnas comienzan en 1
                    palabras = cell.value.split(" ")
                    #print(palabras)
                    if isinstance(cell.value, str) and not cell.value.isdigit() and cell.value.lower()[0] != "z" and "TOTAL" != palabras[0]:
                        #ws1.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=2)
                        for c in row:
                            #Aplica relleno gris a toda la fila
                            c.fill = styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                            c.border = thick_border  # Aplica el borde grueso
                            
                    elif "TOTAL" == palabras[0]:
                        #ws1.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=2)
                        for c in row:
                            #Aplica relleno amarillo a toda la fila
                            c.fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            c.border = thick_border  # Aplica el borde grueso
                            
                
                
                
    # Condición adicional para la última fila, pinta de rojo
    max_row = ws1.max_row
    #ws1.merge_cells(start_row=max_row, start_column=1, end_row=max_row, end_column=2)
    for cell in ws1[max_row]:
        cell.fill = styles.PatternFill(start_color="F15F5F", end_color="F15F5F", fill_type="solid")
        cell.border = thick_border  # Aplica el borde grueso

    detalle = time.strftime("%d-%m-%Y %H hrs %M mins %S segs", time.localtime())
    name = str(detalle) + ' Incentivos_Salida.xlsx'
    wb1.save(name)
    wb1.close()

    # Ruta de destino al Escritorio - Ajusta esto según tu sistema operativo y nombre de usuario
    # Windows: "C:\\Users\\NombreDeUsuario\\Desktop\\"
    ruta_destino = "C:\\Users\\HildaA\\OneDrive - Industrias Tajín\\Desktop\\INCENTIVOS\\"

    # Mover el archivo
    try:
        shutil.move(name, os.path.join(ruta_destino, name))
        print(f"{name} ha sido movido al Escritorio.")
    except FileNotFoundError:
        print(f"No se encontró {name} en la ubicación actual.")
    except Exception as e:
        print(f"Error al mover {name}: {e}")
#########################################################################################################################

# Llamar a la función para seleccionar un archivo
seleccionar_archivo()

# Obtener la ruta del archivo seleccionado
ruta_archivo_excel = os.path.join(os.path.dirname(os.path.abspath(__file__)), "archivo_incentivos.xlsx")

# Cargar el libro de Excel
wb = load_workbook(ruta_archivo_excel)

# Obtener la hoja de trabajo activa
hoja_activa = wb.active

filtro, inicio_sku = filtrar_datos(hoja_activa)
wb.save("archivo_incentivos.xlsx")


# Crear un nuevo libro de trabajo (Workbook)
wb = Workbook()

# Seleccionar la hoja de trabajo activa
hoja_activa = wb.active

#MANIPULACION // CALCULO DE %
for i in range(4, len(filtro)):
    #AQUI CAMBIA LA COLUMNA SI ES DE 2 O EN 1 (el inicio)
    for j in range(inicio_sku, len(filtro[4]), 3):
        filtro[i][j], filtro[i][j+1], filtro[i][j+2] = calcular_porcentaje(filtro[i][j], filtro[i][j+1])
        
#AQUI SE TIENE LA FILA
for i in range(len(filtro)):
    hoja_activa.append(filtro[i])

    
# Guardar el archivo Excel
wb.save("datos.xlsx")

#Bloque de ordenamiento por familias
datos = pd.read_excel("datos.xlsx") # Lee el archivo de excel y lo guarda en un dataframe

"""Orden: Polvo, Líquido, Yaya, Sachet, Chile Seco | Total General | Total Polvos, Total Líquido, Total Yaya, Total Sachet, Total Chile Seco"""

familias = datos.columns.tolist() # Extrae todos los títulos de la primera fila del archivo (las familias).
# Las celdas vacías se van a guardar como "Unnamed" ya que todas las columnas deben tener un encabezado, se va a borrar este encabezado más adelante

# Función que va a recibir un nombre de familia (o encabezado) y va a regresar un dataframe por cada familia, 
# de este modo se van a ir concatenando a un nuevo dataframe conforme al orden predefinido
def ordenar_columnas(familia):
    if familia in familias: # Primero valida que el nombre dado esté en los encabezados del archivo
        columna_inicio = None
        columna_final = None
        contador_columna = 0
        for nombre in familias: # Recorre la lista de encabezados del archivo
            if columna_inicio is None:  # Solo hace esta verificación si aún no se sabe dónde empieza el campo de la familia dada
                if nombre == familia: # Cuando encuentre el nombre dado, toma ese número de columna como inicio del campo de esa familia
                    columna_inicio = contador_columna
            else: # Si ya está definido dónde empieza, ahora va a buscar dónde termina
                if "Unnamed" not in nombre: # Descarta todas las columnas con título "Unnamed" (en el archivo son celdas vacías, aquí se va a agrupar el título de la familia)
                    columna_final = contador_columna # Cuando encuentre el siguiente encabezado, define el final del campo y sale del ciclo
                    break
            contador_columna += 1
    else: return # Si no encuentra el nombre dado en las columnas del archivo, termina la función sin regresar nada
            
    return datos.iloc[:, columna_inicio:columna_final] # Regresa solo una sección del dataframe, correspondiente a las columnas que componen a la familia

contador_columnas_zonas = 0 # Contador para validar si la zona se especifica en una o más columnas
for familia in familias:
    if "Unnamed" not in familia: break
    contador_columnas_zonas += 1
datos_ordenados = datos.iloc[:, :contador_columnas_zonas] # Genera un dataframe que guardará los datos con las columnas ordenadas, comienza con las de zona

# Orden dado, en el orden en el que se definen aquí será el orden en el que se irán enviando a la función y posteriormente al dataframe que se mandará al archivo
orden_familias = ("Polvo", "Líquido", "Yaya", "Sachet", "Chile Seco", "Total General", "Polvo Total", "Líquido Total", "Total Yaya", "Sachet Total", "Chile Seco Total")

# Recorre la lista en la que se define el orden de las familias y manda el nombre a la función, concatena el resultado al dataframe ordenado que ya tiene zonas
for familia in orden_familias:
    nueva_familia = ordenar_columnas(familia)
    datos_ordenados = pd.concat([datos_ordenados, nueva_familia], axis=1)
    
datos_ordenados.to_excel("datos.xlsx", index=False) # Guarda el dataframe ordenado

 # No se puede guardar el dataframe en un excel sin que tenga los encabezados de "Unnamed", por lo que después de guardaar el archivo,
# se abre para leer la primera fila y borrar esos encabezados, para posteriormente mandar el archivo ya limpio al código que combina celdas y da formato
wb1 = load_workbook(filename="datos.xlsx") 
sheets = wb1.sheetnames
sheet = sheets[0]
ws1 = wb1[sheet] 

for row in ws1: 
    for cell in row:
        if cell.row == 1:
            if "Unnamed" in cell.value: cell.value = None # Borra el valor de todas las celdas de la primera fila que tengan valor "Unnamed"
        else: pass

wb1.save('datos.xlsx')
wb1.close()


color_range = 0     # En caso de que se quiera poner de color rojo algún rango entre 0 y algún valor, se define aquí

# Color para cada familia de encabezado
sachet_color = "8ADA39"
polvo_color = "049C04"
liquido_color = "FF0000"
chileseco_color = "D35400"
yaya_color = "454545"
total_color = "FFFF00"

# Diccionario de colores, manualmente puse las keys conforme a los encabezados del archivo
# ChileSeco lo puse junto en caso de que se agregue otra familia cuya primera palabra sea chile
family_colors = {"Sachet": sachet_color, "Polvo": polvo_color, "Yaya": yaya_color, "Líquido": liquido_color, "ChileSeco": chileseco_color, "Total": total_color}
format_file("datos.xlsx", color_range, family_colors) # Llamada a la función principal

# Nombres de los archivos a eliminar
archivo1 = "archivo_incentivos.xlsx"
archivo2 = "datos.xlsx"

# Eliminar archivo1
os.remove(archivo1)
os.remove(archivo2)